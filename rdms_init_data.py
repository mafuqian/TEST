# -*- coding: utf-8 -*-
import os
import time
from openpyxl import load_workbook
import uuid
import argparse
import hashlib

def gen_init_data(xlsx_path, output_file, db='oracle'):
    xlsx_data = load_elsx_data(xlsx_path)
    with open(output_file, 'w') as f:

        f.write("update GP_BM_ID_FILEDATA set DEPART_ID = '{}' where DEPART_ID is not null;\n".format(xlsx_data['info']['TOP_ORG_ID']))
        f.write("update GP_BM_ROLE_INFO set ORG_ID = '{}' where ROLE_NAME = '管理岗_admin';\n".format(xlsx_data['info']['TOP_ORG_ID']))

        f.write("delete from GP_BM_SYS_STAT where SYSTEM_NAME = '{}';\n".format(xlsx_data['info']['MODULE']))
        sysstat = {}
        sysstat['DATA_ID'] = md5_str(xlsx_data['info']['TOP_ORG_ID']+xlsx_data['info']['MODULE'])
        sysstat['CORP_ID'] = xlsx_data['info']['CORP_ID']
        sysstat['ORG_ID'] = xlsx_data['info']['TOP_ORG_ID']
        sysstat['SYSTEM_NAME'] = xlsx_data['info']['MODULE']
        sysstat['SYS_DATE'] =  xlsx_data['info']['RPT_DATE']
        sysstat['LAST_WORK_DATE'] =  xlsx_data['info']['RPT_DATE']
        sysstat['WORK_DATE'] =  xlsx_data['info']['RPT_DATE']
        sysstat['BH_DATE'] =  xlsx_data['info']['RPT_DATE']
        sysstat['SYSTEM_NAME'] = xlsx_data['info']['MODULE']
        sysstat['REPORT_DATE'] = xlsx_data['info']['RPT_DATE']
        sysstat['CURR_REPORT_DATE'] = xlsx_data['info']['RPT_DATE']
        sysstat['STATUS'] = '0'
        add_default_cols(sysstat, True)
        sql = gen_sql('GP_BM_SYS_STAT', sysstat, db)
        f.write(sql+";\n")

        f.write("delete from GP_BM_BRANCH;\n")
        for row in xlsx_data['orgs']:
            sql = gen_sql('GP_BM_BRANCH', row, db)
            f.write(sql+";\n")
        
        f.write("delete from GP_BM_ROLE_INFO WHERE ROLE_ID!='31aa52b7fdb34749969bce5673abab7d';\n")
        for row in xlsx_data['roles']:
            if row['ROLE_ID'] == '31aa52b7fdb34749969bce5673abab7d':
                f.write("delete from GP_BM_ROLE_INFO WHERE ROLE_ID='31aa52b7fdb34749969bce5673abab7d';\n")

            sql = gen_sql('GP_BM_ROLE_INFO', row, db)
            f.write(sql+";\n")
        
        f.write("delete from GP_BM_ROLE_FUNC_REL WHERE ROLE_ID!='31aa52b7fdb34749969bce5673abab7d';\n")
        for row in xlsx_data['role_menus']:
            sql = gen_sql('GP_BM_ROLE_FUNC_REL', row, db)
            f.write(sql+";\n")
        
        f.write("delete from GP_BM_TLR_INFO;\n")
        for row in xlsx_data['users']:
            sql = gen_sql('GP_BM_TLR_INFO', row, db)
            f.write(sql+";\n")
        
        f.write("delete from GP_BM_TLR_ROLE_REL;\n")
        for row in xlsx_data['user_roles']:
            sql = gen_sql('GP_BM_TLR_ROLE_REL', row, db)
            f.write(sql+";\n")

        f.write("delete from GP_BM_BUSINESS_LINE;\n")
        for row in xlsx_data['bizline']:
            sql = gen_sql('GP_BM_BUSINESS_LINE', row, db)
            f.write(sql+";\n")

        f.write("delete from CCRS_BM_RPT_DEPART_REL;\n")
        for row in xlsx_data['bizline_rpt']:
            sql = gen_sql('CCRS_BM_RPT_DEPART_REL', row, db)
            f.write(sql+";\n")

        f.write("delete from CCRS_BM_RPT_CBRC_CFG;\n")
        for row in xlsx_data['rpt_orgs']:
            sql = gen_sql('CCRS_BM_RPT_CBRC_CFG', row, db)
            f.write(sql+";\n")
        
        f.write("delete from CCRS_BM_EX_RATE where DATA_RPT_DATE='{}';\n".format(xlsx_data['info']['RPT_DATE']))
        for row in xlsx_data['ex_rate']:
            sql = gen_sql('CCRS_BM_EX_RATE', row, db)
            f.write(sql+";\n")

        

def load_elsx_data(file_path):
    xlsx_data = {}
    wb = load_workbook(filename=file_path)
    ws_info = wb['基本信息']
    xlsx_data['info'] = {}
    xlsx_data['info']['tpl_version'] = ws_info['B1'].value
    xlsx_data['info']['CORP_ID'] = ws_info['B2'].value
    xlsx_data['info']['CORP_NAME'] = ws_info['B3'].value
    xlsx_data['info']['TOP_ORG_ID'] = ws_info['B4'].value
    xlsx_data['info']['MODULE'] = ws_info['B5'].value
    xlsx_data['info']['RPT_DATE'] = ws_info['B6'].value

    ws_org = wb['机构']
    xlsx_data['orgs'] = []
    for ri in range(3, ws_org.max_row+1):
        org_info = {}
        org_info['DATA_ID'] = md5_str(xlsx_data['info']['TOP_ORG_ID'])
        org_info['DATA_DATE'] = time.strftime("%Y%m%d", time.localtime())
        org_info['CORP_ID'] = xlsx_data['info']['CORP_ID']
        org_info['ORG_ID'] = xlsx_data['info']['TOP_ORG_ID']
        org_info['BRCODE'] = ws_org['A'+str(ri)].value
        org_info['BRNO'] = ws_org['A'+str(ri)].value
        org_info['BRNAME'] = ws_org['B'+str(ri)].value
        org_info['BRCLASS'] = ws_org['C'+str(ri)].value
        org_info['BLN_BRANCH_BRCODE'] = '01'
        org_info['BLN_UP_BRCODE'] = ws_org['D'+str(ri)].value
        org_info['STATUS'] = '1'
        org_info['ST'] = '4'
        org_info['IS_LOCK'] = '0'
        org_info['IS_DEL'] = 'F'
        org_info['GPMS_NEXTACTION'] = '21' 
        org_info['BUSINESS_LINE'] = ws_org['E'+str(ri)].value
        add_default_cols(org_info)
        xlsx_data['orgs'].append(org_info)

    ws_role = wb['角色']
    xlsx_data['roles'] = []
    for ri in range(3, ws_role.max_row+1):
        role_info = {}
        role_info['DATA_ID'] = ws_role['A'+str(ri)].value
        role_info['CORP_ID'] = xlsx_data['info']['CORP_ID']
        role_info['ROLE_ID'] = ws_role['A'+str(ri)].value
        role_info['ROLE_NAME'] = ws_role['B'+str(ri)].value
        role_info['STATUS'] = '1'
        role_info['BUSINESS_LINE'] =  ws_role['C'+str(ri)].value
        add_default_cols(role_info, True)
        xlsx_data['roles'].append(role_info)

    ws_user = wb['用户']
    xlsx_data['users'] = []
    xlsx_data['user_roles'] = []
    for ri in range(3, ws_user.max_row+1):
        user_info = {}
        user_info['DATA_ID'] = ws_user['A'+str(ri)].value
        user_info['CORP_ID'] = xlsx_data['info']['CORP_ID']
        user_info['ORG_ID'] =  xlsx_data['info']['TOP_ORG_ID']
        user_info['GROUP_ID'] = 'XD'
        user_info['TLRNO'] = ws_user['A'+str(ri)].value
        user_info['TLR_NAME'] = ws_user['B'+str(ri)].value
        user_info['PASSWORD'] = '$2a$10$dwwKbUcFnqlaYTZy9BbE5uzWzdEgfgy4KPVXKqMCiUIFindJeLhRa'
        user_info['PASSWD_ENC'] = 'bcrypt'
        user_info['BRCODE'] = ws_user['C'+str(ri)].value
        user_info['BRNO'] = ws_user['C'+str(ri)].value
        user_info['STATUS'] = '0'
        user_info['ST'] = '4'
        user_info['IS_LOCK'] = '0'
        user_info['PSWD_ERR_CNT'] = '0'
        user_info['TOT_PSWD_ERR_CNT'] = '0'
        user_info['FLAG'] = '1'
        user_info['CREDATE_DATE'] = time.strftime("%Y%m%d", time.localtime())
        user_info['LAST_UPD_OPR_ID'] = 'admin'
        user_info['LAST_UPD_TIME'] = time.strftime("%Y%m%d%H%M%S", time.localtime())
        user_info['EMAIL'] =  ws_user['D'+str(ri)].value if ws_user['D'+str(ri)].value is not None  else ''
        user_info['GPMS_NEXTACTION'] = '21'

        add_default_cols(user_info, True)
        xlsx_data['users'].append(user_info)

        if ws_user['E'+str(ri)].value is  None:
            continue

        rolelist =  ws_user['E'+str(ri)].value
        user_roles = rolelist.split(',')
        for ur in user_roles:
            user_role = {}
            user_role['DATA_ID'] = user_info['TLRNO']+'-'+ur.strip()
            user_role['CORP_ID'] = xlsx_data['info']['CORP_ID']
            user_role['TLRNO'] = user_info['TLRNO']
            user_role['ROLE_ID'] = ur.strip()
            user_role['STATUS'] = '1'
            add_default_cols(user_role, True)
            xlsx_data['user_roles'].append(user_role)

    ws_bizline = wb['业务线']
    xlsx_data['bizline'] = []
    xlsx_data['bizline_rpt'] = []
    for ri in range(3, ws_bizline.max_row+1):
        bizline = {}
        bizline['DATA_ID'] = ws_bizline['A'+str(ri)].value
        bizline['CORP_ID'] = xlsx_data['info']['CORP_ID']
        bizline['ORG_ID'] = xlsx_data['info']['TOP_ORG_ID']
        #bizline['GROUP_ID'] = ws_bizline['A'+str(ri)].value
        bizline['BUSINESS_LINE'] = ws_bizline['A'+str(ri)].value
        bizline['BUSINESS_LINE_NAME'] = ws_bizline['B'+str(ri)].value
        bizline['STATUS'] = '1'
        bizline['GPMS_NEXT_ACTION'] = '11'
        add_default_cols(bizline, True)
        xlsx_data['bizline'].append(bizline)

        if ws_bizline['C'+str(ri)].value is None:
            continue

        rptlist =  ws_bizline['C'+str(ri)].value
        bizline_rpts = rptlist.split(',')
        for rpt in bizline_rpts:
            bizline_rpt = {}
            bizline_rpt['DATA_ID'] = bizline['BUSINESS_LINE']+'-'+rpt.strip()
            bizline_rpt['CORP_ID'] = xlsx_data['info']['CORP_ID']
            bizline_rpt['REPORT_CODE'] = rpt.strip()
            bizline_rpt['REPORT_NAME'] = ''
            bizline_rpt['BUSINESS_LINE'] = bizline['BUSINESS_LINE']
            bizline_rpt['BUSINESS_LINE_NAME'] =  bizline['BUSINESS_LINE_NAME']
            add_default_cols(bizline_rpt, True)
            xlsx_data['bizline_rpt'].append(bizline_rpt)

    ws_rpt_org = wb['上报行']
    xlsx_data['rpt_orgs'] = []
    for ri in range(3, ws_rpt_org.max_row+1):
        rpt_org = {}
        rpt_org['DATA_ID'] = ws_rpt_org['A'+str(ri)].value
        rpt_org['CORP_ID'] = xlsx_data['info']['CORP_ID']
        rpt_org['ORG_ID'] =  xlsx_data['info']['TOP_ORG_ID']
        rpt_org['NBJGH'] = ws_rpt_org['A'+str(ri)].value
        rpt_org['JRXKZH'] = ws_rpt_org['B'+str(ri)].value
        rpt_org['YXJGMC'] = ws_rpt_org['C'+str(ri)].value
        rpt_org['IS_REPORT'] = ws_rpt_org['D'+str(ri)].value
        #add_default_cols(rpt_org, True)
        rpt_org['DATA_DATE'] = xlsx_data['info']['RPT_DATE']
        xlsx_data['rpt_orgs'].append(rpt_org)

    ws_ex_rate = wb['汇率']
    xlsx_data['ex_rate'] = []
    for ri in range(3, ws_ex_rate.max_row+1):
        ex_rate = {}
        ex_rate['DATA_ID'] =  xlsx_data['info']['RPT_DATE']+'-'+ws_ex_rate['A'+str(ri)].value
        ex_rate['CORP_ID'] = xlsx_data['info']['CORP_ID']
        ex_rate['DATA_RPT_DATE'] = xlsx_data['info']['RPT_DATE']
        ex_rate['CODE'] = ws_ex_rate['A'+str(ri)].value
        ex_rate['NAME'] = ws_ex_rate['B'+str(ri)].value
        ex_rate['UNIT'] = '1'
        ex_rate['TO_USD'] = str(ws_ex_rate['C'+str(ri)].value)
        add_default_cols(ex_rate, True)
        xlsx_data['ex_rate'].append(ex_rate)

    ws_menus = wb['功能列表']
    xlsx_data['role_menus'] = []
    COLS_LABEL=['A','B','C','D','E','F','G','H','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    role_ids = {}
    for ci in range(4, ws_menus.max_column+1):
        if ws_menus[COLS_LABEL[ci-1]+'2'].value is  None:
            break

        role_ids[COLS_LABEL[ci-1]] =  ws_menus[COLS_LABEL[ci-1]+'2'].value

    for ri in range(3, ws_menus.max_row+1):
        for ci in range(4, ws_menus.max_column+1):
            col_name = COLS_LABEL[ci-1]
            if not role_ids.__contains__(col_name):
                break
            if ws_menus[col_name+str(ri)].value is None:
                continue

            if ws_menus[col_name+str(ri)].value != 'Y':
                continue

            rptlist =  ws_menus[col_name+str(ri)]
            role_menu = {}
            #role_menu['ID'] = role_ids[col_name]+'-'+ws_menus['B'+str(ri)].value
            role_menu['ID'] = md5_str(role_ids[col_name]+'-'+ws_menus['B'+str(ri)].value)
            #print(role_menu['ID'])
            #role_menu['CORP_ID'] = xlsx_data['info']['CORP_ID']
            role_menu['ROLE_ID'] = role_ids[col_name]
            role_menu['FUNCID'] = ws_menus['B'+str(ri)].value
            xlsx_data['role_menus'].append(role_menu)


    return xlsx_data
   

def add_default_cols(cols, all=False):
    if all:
        #cols['DATA_ID'] = uuid.uuid1()
        cols['DATA_DATE'] =  time.strftime("%Y%m%d", time.localtime())
        cols['DATA_SOURCE'] = 'O'
        cols['CHECK_FLAG'] = 'N'
        cols['DATA_VERSION'] = '0'
       

    cols['NEXT_ACTION'] = '99'
    cols['DATA_STATUS'] = '04'
    cols['DATA_FLAG'] = '0'
    cols['DATA_CRT_USER'] = 'admin' 
    cols['DATA_CRT_DATE'] = time.strftime("%Y%m%d", time.localtime())
    cols['DATA_CRT_TIME'] = time.strftime("%Y%m%d%H%M%S", time.localtime())
    #cols['DATA_CHG_USER'] = 
    #cols['DATA_CHG_DATE'] = 
    #cols['DATA_CHG_TIME'] = 
    #cols['DATA_APV_USER'] = 
    #cols['DATA_APV_DATE'] = 
    #cols['DATA_APV_TIME'] = 
    											
def gen_sql(table_name, cols, db='oracle'):
    cols_str = ''
    vals_str = ''

    for k,v in cols.items():
        if len(cols_str) > 0:
            cols_str = cols_str + ','
            vals_str = vals_str + ','  
        
        
        if db == 'mysql':
            cols_str = cols_str+'`{}`'.format(k)
        else:
            cols_str = cols_str+'{}'.format(k)

        vals_str = vals_str+"'{}'".format(v)

    sql = 'insert into {}({}) values({})'.format(table_name, cols_str, vals_str)

    return sql

def md5_str(src):
    return hashlib.md5(src.encode(encoding='UTF-8')).hexdigest()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="RDMS init data generator.")
    parser.add_argument("-D", "--db", choices=["mysql", "oracle",'sqlserver','sybase'], default="oracle", help="db type:mysql oracle sqlserver sybase")

    args = parser.parse_args()
    print('db type:{}'.format(args.db))
    dist_file_path = 'E:\code_project\data\\demo_base_data\init_'+time.strftime("%Y%m%d", time.localtime())+'.sql'
    gen_init_data('E:\code_project\data\\demo_base_data\RDMS系统初始化.xlsx', dist_file_path, args.db)
    print("output file:"+dist_file_path)